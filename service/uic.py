import psycopg2
from openpyxl import load_workbook
import os
from database import user, dbs, passw, server, port
from modules.autenticar import get_password_hash


async def actualizar_pass_hash():
    conn = psycopg2.connect(dbname=dbs, user=user, password=passw, host=server, port=port)
    cursor = conn.cursor()
    try:
        cursor.execute(f"SELECT ci FROM usuarios where desac=TRUE")
        for fila in cursor.fetchall():
            ci = fila[0]
            nuevo_ci = await get_password_hash(ci)
            sql = "UPDATE usuarios SET hash_clave = %s WHERE ci = %s"
            cursor.execute(sql, (nuevo_ci, ci))
            print(ci)
            conn.commit()
        print("Todos los datos en la tabla usuario ha sido actualizado")
    except psycopg2.Error as e:
        conn.rollback()
        print("Error al cambiar datos de la tabla:", e)
    finally:
        cursor.close()
        conn.close()


async def vinculacion():
    conn = psycopg2.connect(dbname=dbs, user=user, password=passw, host=server, port=port)
    cursor = conn.cursor()

    ruta_archivo_excel = os.getenv('VINCULACION')
    hoja_trabajo = load_workbook(filename=ruta_archivo_excel)['Todo']

    for fila in hoja_trabajo.iter_rows(min_row=2, values_only=True):
        bodega, total_nucleo, tienda, cadena, municipio, provincia = fila

        try:
            consulta = "SELECT id_municipio FROM municipios, provincias WHERE provincias.id_provincia = municipios.id_provincia and LOWER(provincias.nombre) LIKE LOWER(%s) and LOWER(municipios.nombre) LIKE LOWER(%s) LIMIT 1; "
            cursor.execute(consulta, ('%' + provincia + '%', '%' + municipio + '%'))
            municipio = cursor.fetchone()

            consulta = "SELECT id_cadena FROM cadenas WHERE LOWER(cadenas.siglas) LIKE LOWER(%s) LIMIT 1; "
            cursor.execute(consulta, ('%' + cadena + '%',))
            cadena = cursor.fetchone()

            if municipio and cadena:
                municipio = str(municipio[0])
                cadena = str(cadena[0])

                consulta = "SELECT id_tienda FROM tiendas WHERE id_municipio = %s and LOWER(nombre) LIKE LOWER(%s) LIMIT 1;"
                cursor.execute(consulta, (municipio, '%' + tienda + '%'))
                rtienda = cursor.fetchone()
                if not rtienda:
                    consulta = ("INSERT INTO tiendas (nombre, direccion, desac, frecuencia_venta, id_municipio, id_cadena) "
                                "VALUES (%s, %s, %s, %s, %s, %s);")
                    cursor.execute(consulta, (tienda, tienda, False, 0, municipio, cadena))
                    conn.commit()
                    cursor.execute("SELECT id_tienda FROM tiendas WHERE nombre LIKE %s and id_municipio = %s and id_cadena = %s LIMIT 1;")
                    cursor.execute(consulta, ('%' + tienda + '%', municipio, cadena))
                    tienda = cursor.fetchone()
                tienda = rtienda[0]

                consulta = "SELECT id_bodega FROM bodegas, oficinas WHERE bodegas.id_oficina = oficinas.id_oficina and id_municipio = %s and LOWER(numero) LIKE LOWER(%s) LIMIT 1;"
                cursor.execute(consulta, (municipio, bodega))
                bodega = cursor.fetchone()

                if bodega:
                    sql = "UPDATE bodegas SET id_tienda = %s WHERE id_bodega = %s"
                    cursor.execute(sql, (tienda, bodega))

                    cursor.execute("SELECT id_tienda FROM tiendas WHERE nombre LIKE %s and id_municipio = %s and id_cadena = %s LIMIT 1;")
                    cursor.execute(consulta, ('%' + tienda + '%', municipio, cadena))
                    tienda = cursor.fetchone()
            print("Todos los datos en la tabla usuario ha sido actualizado")
        except (Exception, ):
            print("Error al cambiar datos de la tabla:")
    cursor.close()
    conn.close()